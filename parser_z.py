import openpyxl
import requests
from bs4 import BeautifulSoup
import data
from data import print_data,one_hourse,ret_min,print_data_time 
import os
from config import *
from PIL import Image
import json
# from googletrans import Translator
import sqlite3
import random
import telebot
import config

second_zmina = ["21-Б","22-Б","31-Б","41-Б","42-Б","21-Д","22-Д","31-Д","41-Д","42-Д"]
second_zmina1 = ["21-О","22-О","31-О","21-П","31-П","21-Ф","31-Ф","21-М","31-М"]

def get_chat_ids():
    # Встановлюємо з'єднання з базою даних
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()

    # Виконуємо запит для отримання всіх значень chat_id
    cursor.execute("SELECT user_id FROM users")
    rows = cursor.fetchall()

    # Створюємо список chat_id
    chat_ids = [row[0] for row in rows]

    # Закриваємо з'єднання
    conn.close()

    # Повертаємо список chat_id
    return chat_ids

def parse_all_users(user):
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()

    # Виконуємо запит для отримання всіх значень chat_id
    cursor.execute("SELECT * FROM users WHERE user_id = ?", (user,))
    rows = cursor.fetchall()
    print(rows)
    rows = list(rows[0])
    
    return f"№{rows[0]} - <a href='tg://user?id={rows[6]}'>{rows[2]}</a> - {rows[5]}"

def parse_user_game(user):
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()

        # Виконуємо запит для отримання всіх значень chat_id
        cursor.execute("SELECT * FROM users WHERE user_id = ?", (user,))
        rows = cursor.fetchall()
        rows = list(rows[0])
        if rows[1]:
            profile_link = f"https://telegram.me/{rows[1]}"
        else:
            profile_link = f"tg://user?id={rows[6]}"
        
        return f"<a href='{profile_link}'>{rows[2]}</a>"
    except:
        pass

def users_rating(chat_id):
    # list_name = ["Пушок","Мурчик","Лапуля","Котик","Кроляшка","Пушистик","Хом'ячок","Панда","Мишка","Білочка"]
    # random.shuffle(list_name)
    chat_id = str(chat_id)
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM user')
    existing_row = cursor.fetchall()
    list_u = []
    element = []
    for i in existing_row:
        element = list(i)
        element[1] = len(i[1].split("~"))
        list_u.append(element)
    sorted_list = sorted(list_u, key=lambda x: x[1], reverse=True)
    for index, item in enumerate(sorted_list):
        if chat_id in item:
            your_top = index+1
            break
    else:
        return "Ти ще не брав участь у грі."
    if your_top <=10:
        return_list = []
        for i in range(0,10):
            if sorted_list[i][0] == chat_id:
                return_list.append(f"{i+1}) Ти😎 - {sorted_list[i][1]}")
            else:
                return_list.append(f"{i+1}) {parse_user_game(sorted_list[i][0])} - {sorted_list[i][1]}")
        
        return "\n".join(return_list)
    else:
        return_list = []
        for i in range(0,10):
            return_list.append(f"{i+1}) {parse_user_game(sorted_list[i][0])} - {sorted_list[i][1]}")
        return_list.append(f"{your_top}) Ти😎 - {emount_answer(chat_id)}")    
        return "\n".join(return_list)


def emount_answer(chat_id):
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM user WHERE user_id = ?', (chat_id,))
    existing_row = cursor.fetchone()
    if existing_row:
        user_id, list_correct_question, amount_health, time = existing_row
        if len(list_correct_question.split("~"))>1:
            return len(list_correct_question.split("~"))
        elif list_correct_question:
            return 1
        else:
            return 0

    return 0

def return_minets(chat_id):
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM user WHERE user_id = ?', (chat_id,))
    existing_row = cursor.fetchone()
    user_id, list_correct_question, amount_health, time = existing_row
    return ret_min(time)


def add_health(chat_id):
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM user WHERE user_id = ?', (chat_id,))
    existing_row = cursor.fetchone()
    user_id, list_correct_question, amount_health, time = existing_row
    if one_hourse(time):
        amount_health = "10"
        cursor.execute('UPDATE user SET list_correct_question = ?, amount_health = ?, data_time_answer = ? WHERE user_id = ?',(list_correct_question, amount_health, time, chat_id))
        conn.commit()
        conn.close()
        return True
    return False
    

def user_health(chat_id):
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS user (
                    user_id TEXT,
                    list_correct_question TEXT,
                    amount_health TEXT,
                    data_time_answer TEXT
                )''')   
    cursor.execute('SELECT * FROM user WHERE user_id = ?', (chat_id,))
    existing_row = cursor.fetchone()
    if existing_row:
        user_id, list_correct_question, amount_health, data_time_answer = existing_row
        return int(amount_health)
    return 1

def users_db(chat_id, correct, id_question):
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM user WHERE user_id = ?', (chat_id,))
    existing_row = cursor.fetchone()
    time = print_data_time()
    if existing_row:
        user_id, list_correct_question, amount_health, data_time_answer = existing_row
        if correct:
            if int(amount_health)>0:
                if len(list_correct_question)>0:
                    list_correct_question += f"~{id_question}"
                else:
                    list_correct_question = str(id_question)
                cursor.execute('UPDATE user SET list_correct_question = ?, amount_health = ?, data_time_answer = ? WHERE user_id = ?',(list_correct_question, amount_health, data_time_answer, user_id))

            else:
                return "error_health"
        else:
            if int(amount_health)>0:
                amount_health = str(int(amount_health)-1)
                if int(amount_health)==0:
                    data_time_answer = time
                cursor.execute('UPDATE user SET list_correct_question = ?, amount_health = ?, data_time_answer = ? WHERE user_id = ?',(list_correct_question, amount_health, data_time_answer, user_id))

            else:
                return "error_health"
        # Рядок з вказаним user_id вже існує, замінюємо дані
    else:
        # Рядок з вказаним user_id не знайдено, додаємо новий рядок
        if correct:
            cursor.execute('INSERT INTO user (user_id, list_correct_question, amount_health, data_time_answer) VALUES (?, ?, ?, ?)', (chat_id, id_question, "10", time))
        else:
            cursor.execute('INSERT INTO user (user_id, list_correct_question, amount_health, data_time_answer) VALUES (?, ?, ?, ?)', (chat_id, "", "9", time))

    conn.commit()
    conn.close()

def random_question(chat_id):
    conn = sqlite3.connect('user_db.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM user WHERE user_id = ?', (chat_id,))
    existing_row = cursor.fetchone()

    conn.commit()
    conn.close()


    conn = sqlite3.connect('database_question.db')
    cursor = conn.cursor()

    # Вибираємо випадковий рядок з таблиці
    cursor.execute("SELECT * FROM my_table ORDER BY RANDOM() LIMIT 1")
    random_row = cursor.fetchone()
    id = random_row[0]
    question = random_row[1]
    answers = random_row[2].split("~")
    correct_answer = random_row[3]
    # Виводимо випадковий рядок
    
    if existing_row:
        user_id, list_correct_question, amount_health, data_time_answer = existing_row
        correct_list = list_correct_question.split("~")
        # print(correct_list)
        if str(id) in correct_list:
            return random_question(chat_id)
    print("quest - ",random_row[0])
    # Закриваємо з'єднання з базою даних
    cursor.close()
    conn.close()
    return id, question, answers, correct_answer

def insert_into_table(arg1, arg2, arg3):
    # Підключення до бази даних
    conn = sqlite3.connect('database_question.db')
    c = conn.cursor()

    # Створення таблиці, якщо вона ще не існує
    c.execute('''CREATE TABLE IF NOT EXISTS my_table (
                    id INTEGER PRIMARY KEY,
                    question TEXT,
                    answers TEXT,
                    correct_answer TEXT
                )''')
    query = "SELECT COUNT(*) FROM my_table WHERE question = ?"
    c.execute(query, (arg1,))
    result = c.fetchone()[0]

    query = "SELECT COUNT(*) FROM my_table"
    c.execute(query)
    amount = c.fetchone()[0]

    if result == 0:
    # Вставка даних до таблиці
        print(f"{amount} - {arg1}")
        c.execute('INSERT INTO my_table (question, answers, correct_answer) VALUES (?, ?, ?)', (arg1, arg2, arg3))
    else:
        print(f" - {arg1}")
    # Збереження змін та закриття підключення до бази даних
    conn.commit()
    conn.close()

# def translate(text):
#     if text is None:
#         return None
#     try:
#         translator = Translator()
#         translation = translator.translate(text, src="en" ,dest='uk')
#         # print(translation.text)
#         return translation.text
#     except Exception as e:
#         print(e)
#         return translate(text)

# def get_random_question():
#     response = requests.get("https://opentdb.com/api.php?amount=50")
#     if response.status_code == 200:
#         question_data = response.json()
#         results = question_data["results"]
#         print(len(results))
#         for n in range(len(results)):
#             if results:
#                 question = results[n]["question"]
#                 answers = results[n]["incorrect_answers"]
#                 ans=[]
#                 for i in answers:
#                     ans.append(translate(i))
#                 answers = ans
#                 correct_answer = translate(results[n]["correct_answer"])
#                 answers.append(correct_answer)
#                 answers.sort()  # перемішуємо відповіді
                
#                 insert_into_table(translate(question), "~".join(answers), correct_answer)
#     return None, None, None



def parse_weather(day):
    city = "Rivne"
    api_key = api_key_pars_weather  # Замініть на свій ключ API
    url = f"http://api.weatherapi.com/v1/forecast.json?key={api_key}&q={city}&days=7&lang=uk"
    response = requests.get(url)
    data = json.loads(response.text)
    forecast_tomorrow = data["forecast"]["forecastday"][day]
    date = forecast_tomorrow["date"]
    max_temp = forecast_tomorrow["day"]["maxtemp_c"]
    min_temp = forecast_tomorrow["day"]["mintemp_c"]
    avg_temp = forecast_tomorrow["day"]["avgtemp_c"]
    condition = forecast_tomorrow["day"]["condition"]["text"]
    image = forecast_tomorrow["day"]["condition"]['icon']
    return(f"Погода в місті {city} на {date}:\nМаксимальна температура: {max_temp} °C \nМінімальна температура: {min_temp} °C \nСередня температура: {avg_temp} °C \nУмови: {condition}" )

def parse_z():
    # try:
    #     file_name = 'zaminu/'+print_data()+'.png'
    #     api_key = api_key_pars_foto
    #     if int(print_data()[:2])%2 == 1:
    #         url = nubip+"/"
    #     else:
    #         url = nubip
    #     script = """
    #         function scrollToBottom() {
    #             window.scrollTo(0, document.body.scrollHeight);
    #         }
    #         scrollToBottom();
    #     """
    #     url = f'https://api.apiflash.com/v1/urltoimage?access_key={api_key}&url={url}&format=png&scroll=true&full_page=true'
    #     response = requests.get(url, params={'js': script})
    #     with open(file_name, 'wb') as f:
    #         f.write(response.content)
    #     img = Image.open(file_name)
    #     # img_cropped = img.crop((400, 250, 1450, img.height-350))
    #     # img_cropped.save(file_name)
    #     print(print_time(),"zaminu")
    #     return True
    # except Exception as e:
    #     print(e)
    #     return False
    try:
        file_name = 'zaminu/'+print_data()+'.png'
        url = "https://rfc.nubip.edu.ua/to-a-student/changes-to-the-schedule/"
        header = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 OPR/107.0.0.0"
        }
        # Виконуємо запит GET до URL
        response = requests.get(url,headers=header)

        # Створюємо об'єкт BeautifulSoup з HTML відповіді
        soup = BeautifulSoup(response.text, 'html.parser')

        # Знаходимо елементи за класом
        elements = soup.find_all(class_="stk-img-wrapper stk-image--shape-stretch stk--shadow-none")
        url = str() 
        # Виводимо URL зображень
        for element in elements:
            img = element.find('img')
            if img and 'data-src' in img.attrs:
                url = img['data-src']
        response = requests.get(url,headers=header)

        # Перевіряємо, чи запит був успішним
        if response.status_code == 200:
            # Зберігаємо зображення
            with open(file_name, 'wb') as f:
                f.write(response.content)
        else:
            print(f"Не вдалося завантажити зображення. Код статусу: {response.status_code}")
        return True
    except Exception as e:
        bot = telebot.TeleBot(config.TOKEN)
        bot.send_message(config.Admin, 'парс замін пройшов не успішно(((((((((((((')
        print(e)
        return False
def ch_z():
    with open("chuselnuk_znamenuk.txt","r", encoding='utf-8') as file:
        if 'чисельник' == file.readline():
            return True
        else:
            return False

def zmina_ch_z():
    with open("chuselnuk_znamenuk.txt","r", encoding='utf-8') as file:
        if 'чисельник' == file.readline():
            with open("chuselnuk_znamenuk.txt","w", encoding='utf-8') as f:
                f.write("знаменик")
        else:
            with open("chuselnuk_znamenuk.txt","w", encoding='utf-8') as f:
                f.write("чисельник")


def parse_r(group:str,day:int,subota = False,teacher = False):
    try:
        week = "ABCDE"
        roz = ["08:30-09:50","10:00-11:20","11:30-12:50","13:00-14:20","14:50-16:10","16:20-17:40","17:50-19:10"]
        roz_subota = ["08:00-09:00","09:10-10:10","10:20-11:20","11:30-12:30","12:40-13:40","13:50-14:50","15:00-16:00"]
        wb = openpyxl.load_workbook('schedule.xlsx')
        sheet = wb[group]
        message = ''
        if not subota:
            for i in range(1,8):
                value = sheet[week[day-1]+str(i)].value
                if "/" in value:
                    if ch_z():
                        value = value[0:value.find("/")]
                    else:
                        value = value[value.find("/")+2:]
                if len(value)>2:
                    values = value.split("\n")
                    message += str(i)+") "+roz[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        else:
            for i in range(1,8):
                value = sheet[week[day-1]+str(i)].value
                if "/" in value:
                    if ch_z():
                        value = value[0:value.find("/")]
                    else:
                        value = value[value.find("/")+2:]
                if len(value)>2:
                    values = value.split("\n")
                    message += str(i)+") "+roz_subota[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        # if teacher:
        #     text = "\n".join(find_elements_by_digit(group,day,ch_z()))
        #     if not text:
        #         text = "Пар немає😇"
        #     return text


        # week = "ABCDE"
        # second_zmina = ["21-Б","22-Б","31-Б","41-Б","42-Б","21-Д","22-Д","31-Д","41-Д","42-Д"]
        # second_zmina1 = ["21-О","22-О","31-О","21-П","31-П","21-Ф","31-Ф","21-М","31-М"]
        
        # roz = ["08:30-09:50","10:00-11:20","11:30-12:50","13:00-14:20"]
        # roz1 = ["13:00-14:20","14:50-16:10","16:20-17:40","17:50-19:10","19:20-20:40","19:20 – 20:40"]
        # roz2 = ["11:30-12:50","13:00-14:20","14:50-16:10","16:20-17:40","17:50-19:10"]
        # roz_subota = ["08:00-09:00","09:10-10:10","10:20-11:20","11:30-12:30"]
        # roz_subota1 = ["11:30-12:30","13:00–14:00","14:10–15:10","15:20–16:20","16:30–17:30"]
        # roz_subota2 = ["10:20-11:20","11:30-12:30","13:00–14:00","14:10–15:10","15:20–16:20"]
        # wb = openpyxl.load_workbook('schedule.xlsx')
        # sheet = wb[group]
        # message = ''
        # if not group in second_zmina and not group in second_zmina1:
        #     if not subota:
        #         for i in range(1,5):
        #             value = sheet[week[day-1]+str(i)].value
        #             if "/" in value:
        #                 if ch_z():
        #                     value = value[0:value.find("/")]
        #                 else:
        #                     value = value[value.find("/")+2:]
        #             if len(value)>2:
        #                 values = value.split("\n")
        #                 message += str(i)+") "+roz[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        #     else:
        #         for i in range(1,5):
        #             value = sheet[week[day-1]+str(i)].value
        #             if "/" in value:
        #                 if ch_z():
        #                     value = value[0:value.find("/")]
        #                 else:
        #                     value = value[value.find("/")+2:]
        #             if len(value)>2:
        #                 values = value.split("\n")
        #                 message += str(i)+") "+roz_subota[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        # elif group in second_zmina:
        #     if not subota:
        #         for i in range(1,6):
        #             value = sheet[week[day-1]+str(i)].value
        #             if "/" in value:
        #                 if ch_z():
        #                     value = value[0:value.find("/")]
        #                 else:
        #                     value = value[value.find("/")+2:]
        #             if len(value)>2:
        #                 values = value.split("\n")
        #                 message += str(i+3)+") "+roz1[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        #     else:
        #         for i in range(1,6):
        #             value = sheet[week[day-1]+str(i)].value
        #             if "/" in value:
        #                 if ch_z():
        #                     value = value[0:value.find("/")]
        #                 else:
        #                     value = value[value.find("/")+2:]
        #             if len(value)>2:
        #                 values = value.split("\n")
        #                 message += str(i+3)+") "+roz_subota1[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        # elif group in second_zmina1:
        #     if not subota:
        #         for i in range(1,6):
        #             value = sheet[week[day-1]+str(i)].value
        #             if "/" in value:
        #                 if ch_z():
        #                     value = value[0:value.find("/")]
        #                 else:
        #                     value = value[value.find("/")+2:]
        #             if len(value)>2:
        #                 values = value.split("\n")
        #                 message += str(i+2)+") "+roz2[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        #     else:
        #         for i in range(1,6):
        #             value = sheet[week[day-1]+str(i)].value
        #             if "/" in value:
        #                 if ch_z():
        #                     value = value[0:value.find("/")]
        #                 else:
        #                     value = value[value.find("/")+2:]
        #             if len(value)>2:
        #                 values = value.split("\n")
        #                 message += str(i+2)+") "+roz_subota2[i-1]+" - "+values[0]+"\n\t"+values[1]+"\n"
        return(message)
    except Exception as e:
        print(e)
        return "Помилка"
def user():
    path = "LOG"
    string = ""
    for filename in os.listdir(path):
        with open(path+"/"+filename) as f:
            for file in f:
                a = file.split("{")
                b = ' '.join(a)
                a = b.split('}')
                b = ' '.join(a)
                a = b.split(':')
                b = ' '.join(a)
                a = b.split(',')
                b = ' '.join(a)
                a = b.split('\'')
                b = ' '.join(a)
                b = b.split()
                try:
                    try:
                        if not b[b.index("chat")+2] in string and not b[b.index("chat")+2][0] == '-':
                            string += b[b.index("chat")+2]+'\t'
                    except:
                        if not b[b.index("chat_id")+1] in string:
                            string += b[b.index("chat_id")+1]+'\t'
                        
                    if b[b.index("username")+1] == "None" and not b[b.index("first_name")+1] in string :
                        string += b[b.index("first_name")+1]+"\n"
                    
                    if not b[b.index("username")+1] in string and not b[b.index("username")+1] == "None":
                        string += "@"+b[b.index("username")+1]+"\n"
                        
                except:
                    pass
    return(string)
def db_replase():
    conn = sqlite3.connect('database_question.db')
    cursor = conn.cursor()

    # Вибірка повторюючихся значень зі стовпця "question"
    cursor.execute('SELECT question, COUNT(*) FROM my_table GROUP BY question HAVING COUNT(*) > 1')
    duplicate_rows = cursor.fetchall()

    # Видалення всіх рядків з повторюючимися значеннями, крім одного
    for row in duplicate_rows:
        question = row[0]
        cursor.execute('DELETE FROM my_table WHERE question = ? AND rowid NOT IN (SELECT min(rowid) FROM my_table GROUP BY question)', (question,))

    # Збереження змін у базі даних
    conn.commit()

    # Закриття з'єднання з базою даних
    conn.close()

def parse_bd_user():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT,
            first_name TEXT,
            last_name TEXT,
            nick_name TEXT,
            groups_name TEXT,
            user_id TEXT
        )
    ''')

    # bot = telebot.TeleBot(TOKEN)
    # with open("Users.txt","r") as file:
    #     for user in file:
    #         try:
    #             user = user[:-1]
    #             chat_member = bot.get_chat_member(chat_id=user, user_id=user)
    #             print(chat_member.user.username," ",chat_member.user.first_name," ",chat_member.user.last_name)
    #             with open (f"Users/{user}/{user}_group.txt") as file:
    #                 group = file.readline()
    #             cursor.execute("INSERT INTO users (username, first_name, last_name, nick_name, groups_name, user_id) VALUES (?, ?, ?, ?, ?, ?)", (chat_member.user.username, chat_member.user.first_name, chat_member.user.last_name, None, group, user))
    #         except Exception as e:
    #             print(f"ban - {user} - {e}")
    conn.commit()
    conn.close()
def link_uaser(message):
    user = message.from_user
    # Create the link to the user's profile
    if user.username:
        profile_link = f"https://telegram.me/{user.username}"
    else:
        profile_link = f"tg://user?id={user.id}"
    return f"<a href='{profile_link}'>{user.first_name}</a>"
def add_or_update_user(user_id, nick_name, groups_name, message):
    try:
        bot = telebot.TeleBot(TOKEN)
        chat_member = bot.get_chat_member(chat_id=user_id, user_id=user_id)
        username, first_name, last_name = chat_member.user.username, chat_member.user.first_name, chat_member.user.last_name
        print(username, first_name, last_name)


        # Встановлюємо з'єднання з базою даних
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()

        # Перевіряємо, чи існує вказаний user_id в таблиці
       
        cursor.execute("SELECT COUNT(*) FROM users WHERE user_id = ?", (user_id,))
        count = cursor.fetchone()[0]
        bot = telebot.TeleBot(TOKEN)
        if count == 0:
            
            bot.send_message(Admin,f"new user - {link_uaser(message)}",parse_mode="HTML")
            # Вставляємо новий запис, якщо user_id не існує
            
            cursor.execute("INSERT INTO users (user_id, username, first_name, last_name, nick_name, groups_name) "
                        "VALUES (?, ?, ?, ?, ?, ?)", (user_id, username, first_name, last_name, nick_name, groups_name))
        else:
            bot.send_message(Admin,f"reg user - {link_uaser(message)}",parse_mode="HTML")
            # Оновлюємо існуючий запис, якщо user_id вже існує
            cursor.execute("UPDATE users SET username = ?, first_name = ?, last_name = ?, nick_name = ?, groups_name = ? "
                        "WHERE user_id = ?", (username, first_name, last_name, nick_name, groups_name, user_id))

        # Зберігаємо зміни і закриваємо з'єднання
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"ban - {user_id} - {e}")


def all_teachers():
    wb = openpyxl.load_workbook("schedule.xlsx")
    data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5, values_only=True):
            sheet_data.append(row)
        data[sheet_name] = sheet_data

    excel_data = data
    teachers = []
    # Вивід даних:
    for sheet_name, sheet_data in excel_data.items():
        print(f"Аркуш: {sheet_name}")
        sheet_data = list(sheet_data)
        for i,row in enumerate(sheet_data):
            for n,element in enumerate(row):
                if element!="-" and element!=None:
                    if "/" in element:
                        for element in element.split("/"):
                            if "-" not in element and "_" not in element:
                                print(i,n,end=" - ")
                                element = element.split("\n")[-1].replace("(", "").replace(")", "")
                                print(element)
                                if element not in teachers:
                                    teachers.append(element)
                    else:
                        print(i,n,end=" - ")
                        element = element.split("\n")[-1].replace("(", "").replace(")", "")
                        print(element)
                        if element not in teachers:
                            teachers.append(element)
    print(len(teachers),teachers)
def find_elements_by_digit(digit, column, ch_zn):
    wb = openpyxl.load_workbook("schedule.xlsx")
    elements = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for i,row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
            if sheet_name in second_zmina:
                i+=3
            elif sheet_name in second_zmina1:
                i+=2
            if row[column - 1] != None and row[column - 1]!="-":
                if "/" in row[column - 1]:
                    if ch_zn:
                        element = row[column - 1].split("/")[0]
                    else:
                        element = row[column - 1].split("/")[1]

                    if "-" not in element and "_" not in element:
                        if digit in element:
                            element = element.split("\n")[-2]
                            elements.append(f"{i+1}) - {sheet_name} {element}")
                else:
                    if digit in row[column - 1]:
                        element = row[column - 1].split("\n")[0]
                        elements.append(f"{i+1}) - {sheet_name} {element}")
    return sorted(elements, key=lambda x: int(x.split(")")[0]))

if __name__ == "__main__":
    # for i in range(360):
    #     get_random_question()
    # print(random_question())
    # db_replase()
    # parse_bd_user()
    # while True:
    #     print(find_elements_by_digit(str(input("> ")),2,True))
    # pass
    # parse_bd_user()
    pass